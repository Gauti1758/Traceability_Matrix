import pandas as pd
import re
import mysql.connector
from db_operations import SECBLTable, truncate_table
from db_connection import create_connection, close_connection
from openpyxl import load_workbook, Workbook
import os

def filter_items(df):
    return df[df['Item Type'].str.upper() != 'FOLDER']

def extract_procedure_code(file_name):
    # Regular expression to match procedure codes <DTTBL_00023_PM_001>
    pattern = r"([A-Z]+_\d+_[A-Z]+_\d+)"
    match = re.search(pattern, file_name)
    if match:
        return match.group(0)
    else:
        return ""

def process_data(df):
    # Filter rows where 'Item Type' is not 'Folder'
    filtered_df = filter_items(df)

    # Extract procedure codes for each file name
    filtered_df['Procedure_Code'] = filtered_df['Name'].apply(extract_procedure_code)

    # Count the number of matched and unmatched procedure coded from SEC library
    total_matched = (filtered_df['Procedure_Code'] != "").sum()
    total_unmatched = (filtered_df['Procedure_Code'] == "").sum()

    return filtered_df[['Name', 'Procedure_Code']], total_matched, total_unmatched

def update_unique_procedure_code(tableName):
    connection = create_connection()
    if connection:
        try:
            cursor = connection.cursor()

            #Fetch all rows ordered by procedure_code and row_id
            select_query = f"""
            SELECT row_id, procedure_code
            FROM {tableName}
            ORDER BY procedure_code, row_id
            """
            cursor.execute(select_query)
            records = cursor.fetchall()

            #Dictionary to track first occurrence of each procedure_code
            first_occurrence_map = {}

            for row in records:
                row_id = row[0]
                procedure_code = row[1]

                # Step 3: Check if the procedure_code has already been updated
                if procedure_code not in first_occurrence_map:
                    # If it's the first occurrence, update unique_procedure_code
                    update_query = f"""
                    UPDATE {tableName}
                    SET unique_procedure_code = %s
                    WHERE row_id = %s
                    """
                    cursor.execute(update_query, (procedure_code, row_id))
                    first_occurrence_map[procedure_code] = True

            connection.commit()
            print("Unique procedure codes updated successfully.")

        except Exception as e:
            print(f"Error updating unique_procedure_code: {e}")

        finally:
            close_connection(connection)


def update_status_table(StatusTable, SECTable, ARISTable):
    connection = create_connection()
    truncate_table(connection, StatusTable)
    
    try:
        # Retrieve data from SEC_DTTBL
        sec_query = f"SELECT file_name, procedure_code, unique_procedure_code FROM {SECTable}"
        sec_cursor = connection.cursor()
        sec_cursor.execute(sec_query)
        sec_records = sec_cursor.fetchall()

        # Retrieve data from ARIS_DTTBL
        aris_query = f"SELECT procedure_name, procedure_code FROM {ARISTable}"
        aris_cursor = connection.cursor()
        aris_cursor.execute(aris_query)
        aris_records = aris_cursor.fetchall()

        # Insert into status table
        insert_status_query = f"""
            INSERT INTO {StatusTable} (SEC_Lib_File_Name, File_Procedure_Code, Unique_Procedure_Code, 
            ARIS_Procedure_Name, ARIS_Procedure_Code, White_List, Black_List)
            VALUES (%s, %s, %s, %s, %s, %s, %s)
        """

        # Process each record from SEC_DTTBL
        for sec_row in sec_records:
            procedure_name = sec_row[0]
            procedure_code = sec_row[1]
            unique_procedure_code = sec_row[2]
            # unique_procedure_code = sec_row[3].strip()  # Strip before matching

            aris_procedure_name = None
            aris_procedure_code = None
            white_list = "Not Pass"
            black_list = "Not Pass"

            # Check unique_procedure_code matches any in ARIS <BL>
            if unique_procedure_code:
                matchdata = None
                for aris_row in aris_records:
                    aris_code = aris_row[1].strip()
                    if unique_procedure_code == aris_code:
                        white_list = "Pass"
                        matchdata = True
                        aris_procedure_name = aris_row[0]
                        aris_procedure_code = aris_code
                        break  # Break after finding the first match

                # Check if procedure_code exists in ARIS_DTTBL
                # for aris_row in aris_records:
                #     aris_code = aris_row[1].strip()  # Strip before matching
                #     if unique_procedure_code == aris_code:
                #         black_list = "Not Pass"
                #         break  # Break after finding a match
                if not matchdata:
                    black_list = "Pass"

            else:
                white_list = ""

            # Insert into Status_DTTBL
            status_data = (
                procedure_name,
                procedure_code,
                unique_procedure_code,
                aris_procedure_name,
                aris_procedure_code,
                white_list,
                black_list
            )
            sec_cursor.execute(insert_status_query, status_data)

        connection.commit()
    except mysql.connector.Error as error:
        print(f"Error: {error}")
    finally:
        sec_cursor.close()
        aris_cursor.close()
        close_connection(connection)

def generate_status_excel(StatusTable):
    connection = create_connection()
    
    try:
        #Fetch data from Status_DTTBL
        query = f"""
            SELECT Row_Id, SEC_Lib_File_Name, File_Procedure_Code, Unique_Procedure_Code, 
                   ARIS_Procedure_Name, ARIS_Procedure_Code, White_List, Black_List 
            FROM {StatusTable}
        """

        cursor = connection.cursor()
        cursor.execute(query)
        result = cursor.fetchall()
        header = ["Row_Id", "SEC_Lib_File_Name", "File_Procedure_Code", "Unique_Procedure_Code", 
                   "ARIS_Procedure_Name", "ARIS_Procedure_Code", "White_List", "Black_List"]

        white_list_data = [header] + [list(row) for row in result]
    
    except mysql.connector.Error as error:
        print(f"Error: {error}")
    finally:
        close_connection(connection)

    return white_list_data


def updateRedList(SECTable, ARISTable):
    connection = create_connection()

    try:
        if connection:
            cursor = connection.cursor()

            aris_query = f"""
            SELECT row_id, procedure_name, procedure_code
            FROM {ARISTable}
            """
            cursor.execute(aris_query)
            aris_records = cursor.fetchall()

            sec_query = f"""
            SELECT file_name, unique_procedure_code 
            FROM {SECTable}
            """
            cursor.execute(sec_query)
            sec_records = cursor.fetchall()

            sec_dict = {record[1]: record[0] for record in sec_records}
            status_data = []

            for aris_row in aris_records:
                row_id = aris_row[0]
                procedure_name = aris_row[1]
                procedure_code = aris_row[2]
                red_list = "Pass"  # Default to 'Pass'

                file_name = ""
                unique_procedure_code = ""

                if procedure_code in sec_dict:
                    file_name = sec_dict[procedure_code]
                    unique_procedure_code = procedure_code
                    red_list = "Not Pass"

                status_data.append([
                    row_id,
                    procedure_name,
                    procedure_code,
                    file_name,
                    unique_procedure_code,
                    red_list
                ])

            update_query = f"""
            UPDATE {ARISTable}
            SET Red_List = %s
            WHERE row_id = %s
            """
            for row in status_data:
                cursor.execute(update_query, (row[5], row[0]))  # Update Red_List in ARIS_DTTBL

            connection.commit()
            print("ARIS_DTTBL updated with Red_List successfully.")

            #Insert the status data into the Excel file
            # updateDTTBLStatusExcel(status_data)

    except Exception as e:
        print(f"Error updating Red_List: {e}")
    finally:
        close_connection(connection)
    return status_data


def updateDTTBLStatusExcel(aris_data):
    try:
        # Define the file path
        file_path = 'Output_Docs/DTTBL_Status.xlsx'
        
        # Load the existing Excel file
        book = load_workbook(file_path)
        
        # Use pd.ExcelWriter with openpyxl engine
        with pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
            sheet = writer.sheets.get('Sheet1')

            if not sheet:
                raise ValueError("Sheet 'Sheet1' not found in the workbook.")

            # to get next available row in the sheet
            next_row = sheet.max_row

            # Insert data into excel
            data_to_insert = pd.DataFrame(aris_data, columns=[
                'Row ID', 'ARIS Procedure Name', 'ARIS Procedure Code', 'SEC File Name', 
                'SEC Procedure Code', 'Red_List'
            ])

            data_to_insert.to_excel(writer, startrow=next_row, index=False, header=True)

        print(f"Data inserted successfully in {file_path}.")
    
    except Exception as e:
        print(f"Error inserting data into Excel: {e}")

#I have to check this function result
#count total number of translated or not translated procedure in ARIS
def procedureCodeOccurance(SECTable, ARISTable, StatusTable):
    connection = create_connection()

    try:
        if connection:
            cursor = connection.cursor()
            #count translated and not translated procedure in ARIS
            Arisquery = f"""
                SELECT procedure_code, COUNT(*) FROM {ARISTable}
                GROUP BY procedure_code
            """
            cursor.execute(Arisquery)
            ARISresults = cursor.fetchall()

            ARIS_procedure = f"""SELECT COUNT(*) FROM {ARISTable}"""
            cursor.execute(ARIS_procedure)
            total_procedure = cursor.fetchone()[0]

            Secquery = f"""
                SELECT COUNT(file_name) as file_name, COUNT(unique_procedure_code) as unique_procedure_code
                from {SECTable}
            """
            cursor.execute(Secquery)
            SECresults = cursor.fetchone()

            cursor.execute(f"SELECT COUNT(*) FROM {StatusTable} WHERE White_List = 'Pass'")
            white_list_count = cursor.fetchone()[0]

            cursor.execute(f"SELECT COUNT(*) FROM {StatusTable} WHERE Black_List = 'Pass'")
            black_list_count = cursor.fetchone()[0]

            cursor.execute(f"SELECT COUNT(*) FROM {ARISTable} WHERE Red_List = 'Pass'")
            red_list_count = cursor.fetchone()[0]
            # connection.commit()
            
            not_translated = 0
            translated = 0
            translated_more_than_two = 0
            totalFileName = SECresults[0]
            totalUniqueProcedure = SECresults[1]

            for row in ARISresults:
                procedure_code = row[0]
                count = row[1]

                if count == 1:
                    not_translated = not_translated + 1
                elif count == 2:
                    translated = translated + 1
                elif count > 2:
                    translated_more_than_two = translated_more_than_two + 1
                    print(f"Procedure code {procedure_code} migrated {count} times in ARIS")
            
            print(f"Total number of files in SEC Library {totalFileName}")
            print(f"Total number of unique procedure in SEC Library {totalUniqueProcedure}")
            print(f"Total procedure migrated in ARIS {total_procedure}")
            print(f"Total number of untranslated procedure migrated ARIS: {not_translated}")
            print(f"Total number of translated procedure migrated ARIS: {translated}")
            print(f"Total number of procedure migrated more than twice in ARIS: {translated_more_than_two}")

    except Exception as e:
        print(f"Error: counting translated procedure code {e}")

    finally:
        close_connection(connection)
    return [totalFileName, totalUniqueProcedure, total_procedure, not_translated, translated, translated_more_than_two], white_list_count, black_list_count, red_list_count

def create_summary_report(SECTable, ARISTable, StatusTable):
    # Create a new Excel workbook
    wb = Workbook()

    # Create the "Summary Results" sheet
    summary_sheet = wb.active
    summary_sheet.title = "Summary Results"

    # Add headers and data for the summary

    List, whiteList, blackList, redList = procedureCodeOccurance(SECTable, ARISTable, StatusTable)
    summary_sheet.append(["NUMERICAL FACTS: (Summary result)"])
    summary_sheet.append(["1. Total number of files in SEC Library:", List[0]])
    summary_sheet.append(["2. Total number of 'documents' or 'process models' in SEC Library:", List[1]])
    summary_sheet.append(["3. Total number of files migrated into ARIS:", f"{List[2]} ({List[4]} x 2 + {List[3]})"])
    summary_sheet.append(["4. Total number of translated procedure migrated in ARIS:", List[4]])
    summary_sheet.append(["5. Total number of untranslated procedure migrated in ARIS:", List[3]])
    summary_sheet.append(["6. Total number of procedure migrated more than twice in ARIS:", List[5]])

    summary_sheet.append([""])
    summary_sheet.append(["SUMMARY of Results:"])
    summary_sheet.append(["7. White List Count =", whiteList])
    summary_sheet.append(["8. Black List Count =", blackList])
    summary_sheet.append(["9. Red List Count =", redList])

    summary_sheet.append([""])
    summary_sheet.append(["ACTIONS:"])
    summary_sheet.append(["10. Migration team will migrate", f"{blackList} (Black List Count) into ARIS"])
    summary_sheet.append(["11. PPMD team requested to investigate", f"{redList} (Red List Count) in SEC Library"])
    summary_sheet.append(["12. PPMD team may consider performing translation on", f"{List[3]} documents"])

    # Create "White List" and "Red List" sheets
    white_list_sheet = wb.create_sheet(title="White_List Black_List")
    red_list_sheet = wb.create_sheet(title="Red_List")

    # Sample data for White_List and Red_List sheets
    white_list_data = generate_status_excel(StatusTable)
    red_list_data = updateRedList(SECTable, ARISTable)

    # Populate White_List sheet
    for row in white_list_data:
        white_list_sheet.append(row)

    # Populate Red_List sheet
    for row in red_list_data:
        red_list_sheet.append(row)

    # Save the Excel workbook to a file
    output_path = 'Output_Docs/Summary_Report_DTTBL.xlsx'
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb.save(output_path)
    print(f"Summary report saved to {output_path}")